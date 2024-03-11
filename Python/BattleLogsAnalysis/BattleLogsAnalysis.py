import json
import openpyxl
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def process_file(file_path):
    file_name = os.path.splitext(os.path.basename(file_path))[0]  # 获取文件名（不含扩展名）

    # 读取JSON数据
    with open(file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)

    # 创建Excel工作簿
    workbook = openpyxl.Workbook()
    heroes_sheet = workbook.active
    heroes_sheet.title = "英雄属性"
    headers = ['BattleHID', 'HID', 'CurHP', 'MaxHP', 'HP Per Unit', 'ATK', 'DEF', 'UnitType', 'MaxUnitCount',
               'SkillIds', 'SkillLevels', 'SLGAttrList', 'BattleProperty', 'Position']
    heroes_sheet.append(headers)

    sorted_attackers = sorted(data['Logs'][0]['Attackers'], key=lambda x: x['BattleHID'])
    sorted_defends = sorted(data['Logs'][0]['Defends'], key=lambda x: x['BattleHID'])
    sorted_heroes = sorted_attackers + sorted_defends
    is_first_two = True

    for hero in sorted_heroes:
        # 在进攻方和防守方插入一个空行
        test = hero['BattleHID'] // 10
        print(test)
        if hero['BattleHID'] // 10 == 2 and is_first_two:
            heroes_sheet.append([])
            is_first_two = False
        hp_per_unit = hero['MaxHP'] / hero['MaxUnitCount'] if hero['MaxUnitCount'] != 0 else 0
        row = [hero['BattleHID'], hero['HID'], hero['CurHP'], hero['MaxHP'], hp_per_unit, hero['ATK'], hero['DEF'],
               hero['UnitType'], hero['MaxUnitCount'], str(hero['SkillIds']), str(hero['SkillLevels']),
               str(hero['SLGAttrList']), str(hero['BattleProperty']), hero['Position']]
        heroes_sheet.append(row)

    battle_hids = {log['BattleHID'] for log in data['Logs'] if 'FromHP' in log and 'ToHP' in log}
    # 然后按照BattleHID的数值进行排序
    sorted_hids = sorted(battle_hids)

    # 预先为每个BattleHID创建一个工作表，以确保它们的顺序
    sheets = {}
    for hid in sorted_hids:
        sheet_name = f"部队{hid}"
        sheets[hid] = workbook.create_sheet(title=sheet_name)

    # 处理日志时，使用相应的工作表：
    for log in data['Logs']:
        if 'FromHP' in log and 'ToHP' in log:
            battle_hid = log['BattleHID']
            hit_logs_sheet = sheets[battle_hid]  # 直接从预先创建的sheets字典中获取对应的sheet
            if not hit_logs_sheet._cells:  # 如果工作表是空的，添加表头
                headers = ['BattleHID', 'FromHP', 'ToHP', 'HP Loss', 'EffectHID', 'SkillId', 'EffectType', 'Dead',
                           'ByBunkerHId', 'EffectTriggerIndex', 'BattleTime']
                hit_logs_sheet.append(headers)
            # 计算损失的HP
            hp_loss = log['FromHP'] - log['ToHP']
            # 检查角色是否死亡
            is_dead = log['ToHP'] == 0
            hit_log = [log['BattleHID'], log['FromHP'], log['ToHP'], hp_loss, log['EffectHID'], log['SkillId'],
                       log['EffectType'], is_dead, log['ByBunkerHId'], log['EffectTriggerIndex'], log['BattleTime']]
            hit_logs_sheet.append(hit_log)

    # 设置字体、居中对齐、边框样式和特殊标记样式
    normal_font = Font(name="等线", size=11)  # 设置字体为等线，大小为11
    header_font = Font(name="等线", size=11, bold=True, color="FFFFFF")  # 表头使用等线，加粗，白色文字
    center_aligned_text = Alignment(horizontal="center", vertical="center")
    thin_border_side = Side(border_style="thin")
    square_border = Border(top=thin_border_side, right=thin_border_side, bottom=thin_border_side, left=thin_border_side)
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    dead_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # DEAD列为TRUE时使用红色背景

    # 应用样式到工作表
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        # 冻结首行
        worksheet.freeze_panes = 'A2'
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = normal_font
                cell.alignment = center_aligned_text
                cell.border = square_border
                if cell.row == 1:  # 应用表头样式
                    cell.font = header_font
                    cell.fill = header_fill
                if cell.column == 8:  # Dead列
                    if cell.value is True:
                        cell.fill = dead_fill

        # 设置列宽
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length * 1.2

        # 设置行高
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = 20

    # 获取原始文件所在的目录
    original_directory = os.path.dirname(file_path)  # 原始文件路径
    # 创建新的目录路径
    new_directory = os.path.join(original_directory, '战报分析')

    # 检查这个新目录是否存在，如果不存在，创建它
    if not os.path.exists(new_directory):
        os.makedirs(new_directory)

    # 确定新的文件名和路径
    new_file_path = os.path.join(new_directory, f'{file_name}.xlsx')

    # 保存Excel文件到新的文件夹
    workbook.save(new_file_path)


def open_file_dialog():
    file_path = filedialog.askopenfilename()
    if file_path:
        process_file(file_path)
        label.config(text=f"处理完成: {file_path}")


root = tk.Tk()
root.title('战报分析工具')
root.geometry('600x400')

open_button = tk.Button(root, text='打开战报文件', command=open_file_dialog)
open_button.pack(expand=True)

label = tk.Label(root, text='请打开一个战报文件')
label.pack(expand=True)

root.mainloop()
